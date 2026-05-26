"""
ReportExporter.py  -  Excel report generator for wfmplan.

Produces a professional multi-sheet Excel workbook containing:
  Sheet 1  - Summary          KPI cards + daily summary table
  Sheet 2  - Hourly Detail    Full per-interval results
  Sheet 3  - Weekly Roster    Day x shift wall-chart + detail
  Sheet 4  - Sensitivity      What-if agent count analysis
  Sheet 5  - How to Use       Methodology notes for the analyst
"""

from __future__ import annotations

from datetime import datetime
from math import ceil
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.styles import (
    Alignment, Border, Font, GradientFill, PatternFill, Side
)
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Colour palette
# ---------------------------------------------------------------------------
DARK_BLUE  = '1F3864'
MID_BLUE   = '2E75B6'
LIGHT_BLUE = 'DEEAF1'
ACCENT     = 'ED7D31'
GREEN      = '70AD47'
YELLOW     = 'FFF2CC'
WHITE      = 'FFFFFF'
LIGHT_GREY = 'F2F2F2'
DARK_GREY  = '404040'


def _font(bold=False, size=11, color='000000', italic=False):
    return Font(name='Arial', bold=bold, size=size, color=color, italic=italic)

def _fill(hex_color):
    return PatternFill('solid', start_color=hex_color, fgColor=hex_color)

def _border(style='thin'):
    side = Side(style=style)
    return Border(left=side, right=side, top=side, bottom=side)

def _center():
    return Alignment(horizontal='center', vertical='center', wrap_text=True)

def _left():
    return Alignment(horizontal='left', vertical='center', wrap_text=True)


# ---------------------------------------------------------------------------
# Exporter
# ---------------------------------------------------------------------------

class ReportExporter:
    """
    Generate a polished Excel report workbook.

    Parameters
    ----------
    batch_results   : pd.DataFrame  - output of BatchOptimizer.run_optimization()
    roster          : pd.DataFrame  - output of WeeklyScheduler.build_roster()   (optional)
    sensitivity_data: list[dict]    - output of Optimizer.sensitivity_analysis() (optional)
    title           : str           - report title shown on Summary sheet
    """

    def __init__(
        self,
        batch_results: pd.DataFrame,
        roster: Optional[pd.DataFrame] = None,
        sensitivity_data: Optional[list] = None,
        title: str = 'Workforce Management Staffing Report',
    ):
        self.results = batch_results.copy()
        self.results['interval_start'] = pd.to_datetime(self.results['interval_start'])
        self.results['interval_end'] = pd.to_datetime(self.results['interval_end'])
        self.roster = roster
        self.sensitivity = sensitivity_data or []
        self.title = title
        self.wb = Workbook()
        self.wb.remove(self.wb.active)  # remove default blank sheet

    # ------------------------------------------------------------------
    # Public entry point
    # ------------------------------------------------------------------

    def export(self, path: str = 'staffing_report.xlsx') -> str:
        """Build all sheets and save workbook.  Returns saved path."""
        self._build_summary()
        self._build_hourly_detail()
        if self.roster is not None and not self.roster.empty:
            self._build_weekly_roster()
        if self.sensitivity:
            self._build_sensitivity()
        self._build_how_to_use()

        self.wb.save(path)
        return path

    # ------------------------------------------------------------------
    # Sheet helpers
    # ------------------------------------------------------------------

    def _header_row(self, ws, row: int, cols: list, bg=MID_BLUE, fg=WHITE, size=11):
        for c, label in enumerate(cols, 1):
            cell = ws.cell(row=row, column=c, value=label)
            cell.font = _font(bold=True, color=fg, size=size)
            cell.fill = _fill(bg)
            cell.border = _border()
            cell.alignment = _center()

    def _title_cell(self, ws, row: int, col: int, value, span: int = 1,
                    bg=DARK_BLUE, fg=WHITE, size=14, bold=True):
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = _font(bold=bold, size=size, color=fg)
        cell.fill = _fill(bg)
        cell.alignment = _center()
        if span > 1:
            ws.merge_cells(
                start_row=row, start_column=col,
                end_row=row, end_column=col + span - 1
            )
        return cell

    def _data_row(self, ws, row: int, values: list, alt: bool = False):
        bg = LIGHT_GREY if alt else WHITE
        for c, v in enumerate(values, 1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.fill = _fill(bg)
            cell.border = _border('hair')
            cell.alignment = _left() if isinstance(v, str) else _center()

    def _set_col_widths(self, ws, widths: list):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def _kpi_card(self, ws, row: int, col: int, label: str, value: str,
                  bg=MID_BLUE, unit: str = ''):
        # Label
        lc = ws.cell(row=row, column=col, value=label)
        lc.font = _font(bold=False, size=9, color=WHITE)
        lc.fill = _fill(bg)
        lc.alignment = _center()
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+1)

        # Value
        vc = ws.cell(row=row+1, column=col, value=f"{value}{unit}")
        vc.font = _font(bold=True, size=18, color=DARK_BLUE)
        vc.fill = _fill(LIGHT_BLUE)
        vc.alignment = _center()
        ws.merge_cells(start_row=row+1, start_column=col, end_row=row+1, end_column=col+1)

    # ------------------------------------------------------------------
    # Sheet 1 – Summary
    # ------------------------------------------------------------------

    def _build_summary(self):
        ws = self.wb.create_sheet('Summary')
        ws.sheet_view.showGridLines = False

        # Title banner
        self._title_cell(ws, 1, 1, self.title, span=14, size=16)
        ws.row_dimensions[1].height = 35
        ts = ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%d %b %Y %H:%M')}")
        ts.font = _font(italic=True, size=9, color='888888')
        ws.merge_cells('A2:N2')

        # KPI section header
        self._title_cell(ws, 4, 1, 'KEY METRICS', span=14, bg=MID_BLUE, size=12)

        # Compute KPIs
        r = self.results
        has_shrink = 'agent_req_shrink' in r.columns
        peak_req = int(r['agent_req_shrink'].max()) if has_shrink else int(r['agent_req'].max())
        total_vol = int(r['exp_vol'].sum())
        avg_occ = f"{r['pred_occupancy'].mean()*100:.1f}" if 'pred_occupancy' in r.columns else 'N/A'
        avg_asa = f"{r['pred_asa'].mean():.1f}" if 'pred_asa' in r.columns else 'N/A'
        r['_hrs'] = r.get('agent_req_shrink', r['agent_req']) * r.get('interval_secs', 3600) / 3600
        total_hrs = f"{r['_hrs'].sum():.0f}"
        n_intervals = len(r)

        kpis = [
            ('Peak Agents Required', peak_req, ''),
            ('Total Contact Volume', f'{total_vol:,}', ''),
            ('Avg Occupancy', avg_occ, '%'),
            ('Avg ASA', avg_asa, 's'),
            ('Total Agent Hours', total_hrs, 'h'),
            ('Intervals Planned', n_intervals, ''),
        ]
        col = 1
        for label, value, unit in kpis:
            self._kpi_card(ws, 5, col, label, str(value), unit=unit)
            col += 2
        ws.row_dimensions[5].height = 20
        ws.row_dimensions[6].height = 30

        # Daily summary table
        self._title_cell(ws, 9, 1, 'DAILY SUMMARY', span=14, bg=MID_BLUE, size=12)
        daily_cols = ['Date', 'Total Volume', 'Avg AHT (s)', 'Peak Agents',
                      'Total Agent Hrs', 'Avg Occupancy %', 'Avg ASA (s)']
        self._header_row(ws, 10, daily_cols)

        r['_date'] = r['interval_start'].dt.date
        r['_agent_hrs'] = r.get('agent_req_shrink', r['agent_req']) * r.get('interval_secs', 3600) / 3600
        daily = r.groupby('_date').agg(
            total_vol=('exp_vol', 'sum'),
            avg_aht=('exp_aht', 'mean'),
            peak_agents=(('agent_req_shrink' if has_shrink else 'agent_req'), 'max'),
            total_hrs=('_agent_hrs', 'sum'),
            avg_occ=('pred_occupancy', 'mean'),
            avg_asa=('pred_asa', 'mean'),
        ).reset_index()

        for i, (_, row) in enumerate(daily.iterrows()):
            vals = [
                str(row['_date']),
                int(row['total_vol']),
                round(row['avg_aht']),
                int(row['peak_agents']),
                round(row['total_hrs'], 1),
                f"{row['avg_occ']*100:.1f}%",
                round(row['avg_asa'], 1),
            ]
            self._data_row(ws, 11 + i, vals, alt=(i % 2 == 1))

        self._set_col_widths(ws, [14, 14, 12, 14, 16, 16, 14])

    # ------------------------------------------------------------------
    # Sheet 2 – Hourly Detail
    # ------------------------------------------------------------------

    def _build_hourly_detail(self):
        ws = self.wb.create_sheet('Hourly Detail')
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = 'A3'

        self._title_cell(ws, 1, 1, 'HOURLY STAFFING DETAIL', span=12, bg=DARK_BLUE, size=14)
        ws.row_dimensions[1].height = 30

        # Dynamic column list based on what's in results
        base_cols = ['interval_start', 'interval_end', 'exp_vol', 'exp_aht']
        metric_cols = [c for c in [
            'interval_secs', 'traffic_intensity', 'agent_req', 'agent_req_shrink',
            'pred_occupancy', 'prob_waiting', 'pred_asa', 'pred_sla_pct'
        ] if c in self.results.columns]

        display_cols = base_cols + metric_cols
        headers = {
            'interval_start': 'Interval Start', 'interval_end': 'Interval End',
            'exp_vol': 'Exp Volume', 'exp_aht': 'AHT (s)',
            'interval_secs': 'Interval (s)', 'traffic_intensity': 'Erlangs (ρ)',
            'agent_req': 'Agents Req', 'agent_req_shrink': 'Agents (+Shrink)',
            'pred_occupancy': 'Occupancy', 'prob_waiting': 'P(Wait)',
            'pred_asa': 'Pred ASA (s)', 'pred_sla_pct': 'Pred SLA %',
        }
        self._header_row(ws, 2, [headers.get(c, c) for c in display_cols])

        for i, (_, row) in enumerate(self.results.iterrows()):
            vals = []
            for c in display_cols:
                v = row.get(c, '')
                if c in ('interval_start', 'interval_end'):
                    v = str(v)[:16]
                elif c == 'pred_occupancy' and pd.notna(v):
                    v = f"{float(v)*100:.1f}%"
                elif c == 'prob_waiting' and pd.notna(v):
                    v = f"{float(v)*100:.1f}%"
                elif pd.notna(v) and isinstance(v, float):
                    v = round(v, 2)
                vals.append(v)
            self._data_row(ws, 3 + i, vals, alt=(i % 2 == 1))

        widths = [18, 18, 12, 10, 12, 12, 12, 16, 12, 10, 12, 12]
        self._set_col_widths(ws, widths[:len(display_cols)])

        # Conditional formatting: highlight peak demand rows
        if 'agent_req_shrink' in self.results.columns:
            peak = int(self.results['agent_req_shrink'].max())
            for i, (_, row) in enumerate(self.results.iterrows()):
                if row.get('agent_req_shrink', 0) >= peak * 0.9:
                    for c in range(1, len(display_cols) + 1):
                        ws.cell(row=3 + i, column=c).fill = _fill('FFF2CC')

    # ------------------------------------------------------------------
    # Sheet 3 – Weekly Roster
    # ------------------------------------------------------------------

    def _build_weekly_roster(self):
        ws = self.wb.create_sheet('Weekly Roster')
        ws.sheet_view.showGridLines = False

        self._title_cell(ws, 1, 1, 'WEEKLY SHIFT ROSTER', span=10, bg=DARK_BLUE, size=14)
        ws.row_dimensions[1].height = 30

        # --- Wall-chart pivot ---
        self._title_cell(ws, 3, 1, 'AGENTS PER SHIFT BY DAY', span=10, bg=MID_BLUE, size=11)

        roster = self.roster.copy()
        roster['date'] = pd.to_datetime(roster['date'])

        pivot = roster.pivot_table(
            index=['date', 'day_of_week'],
            columns='shift_name',
            values='agents_assigned',
            fill_value=0,
        ).reset_index()
        pivot.columns.name = None

        shift_cols = [c for c in pivot.columns if c not in ('date', 'day_of_week')]
        header = ['Date', 'Day'] + shift_cols + ['Total']
        self._header_row(ws, 4, header)

        for i, (_, row) in enumerate(pivot.iterrows()):
            total = sum(int(row[s]) for s in shift_cols)
            vals = [str(row['date'])[:10], row['day_of_week']] + \
                   [int(row[s]) for s in shift_cols] + [total]
            self._data_row(ws, 5 + i, vals, alt=(i % 2 == 1))

        # --- Detailed roster ---
        detail_start = 5 + len(pivot) + 3
        self._title_cell(ws, detail_start, 1, 'SHIFT DETAIL', span=8, bg=MID_BLUE, size=11)
        det_headers = ['Date', 'Day', 'Shift', 'Start', 'End',
                       'Duration (h)', 'Productive (h)', 'Agents', 'Total Hrs']
        self._header_row(ws, detail_start + 1, det_headers)

        roster['day_of_week'] = pd.to_datetime(roster['date']).dt.strftime('%A')
        for i, (_, row) in enumerate(roster.iterrows()):
            vals = [
                str(row['date'])[:10],
                row.get('day_of_week', ''),
                row['shift_name'],
                str(row['shift_start'])[:16],
                str(row['shift_end'])[:16],
                round(float(row.get('duration_hrs', 0)), 1),
                round(float(row.get('productive_hrs', 0)), 1),
                int(row['agents_assigned']),
                round(float(row.get('total_hours', 0)), 1),
            ]
            self._data_row(ws, detail_start + 2 + i, vals, alt=(i % 2 == 1))

        self._set_col_widths(ws, [13, 12, 12, 17, 17, 13, 14, 10, 11])

    # ------------------------------------------------------------------
    # Sheet 4 – Sensitivity
    # ------------------------------------------------------------------

    def _build_sensitivity(self):
        ws = self.wb.create_sheet('Sensitivity')
        ws.sheet_view.showGridLines = False

        self._title_cell(ws, 1, 1, 'SENSITIVITY ANALYSIS  (what-if staffing levels)', span=8, bg=DARK_BLUE, size=14)
        ws.row_dimensions[1].height = 30

        desc = ws.cell(row=3, column=1,
            value='Shows how service metrics change as agent count varies around the optimal level. '
                  'Highlighted row = optimal recommended staffing.')
        desc.font = _font(italic=True, size=10)
        ws.merge_cells('A3:H3')

        has_sla = any('pred_sla_pct' in row for row in self.sensitivity)
        cols = ['Agents', 'Is Optimal', 'Occupancy %', 'P(Wait) %',
                'Pred ASA (s)', 'Agents (+Shrink)']
        if has_sla:
            cols.append('Pred SLA %')
        self._header_row(ws, 5, cols)

        for i, row in enumerate(self.sensitivity):
            vals = [
                row.get('agents'),
                'YES' if row.get('is_optimal') else '',
                row.get('occupancy_pct'),
                row.get('prob_waiting_pct'),
                row.get('pred_asa_sec'),
                row.get('agent_req_shrink'),
            ]
            if has_sla:
                vals.append(row.get('pred_sla_pct', ''))

            for c, v in enumerate(vals, 1):
                cell = ws.cell(row=6 + i, column=c, value=v)
                if row.get('is_optimal'):
                    cell.fill = _fill('E2EFDA')
                    cell.font = _font(bold=True, color='375623')
                else:
                    cell.fill = _fill(WHITE if i % 2 == 0 else LIGHT_GREY)
                cell.border = _border('hair')
                cell.alignment = _center()

        self._set_col_widths(ws, [10, 12, 14, 14, 14, 18, 12])

    # ------------------------------------------------------------------
    # Sheet 5 – How to Use
    # ------------------------------------------------------------------

    def _build_how_to_use(self):
        ws = self.wb.create_sheet('How to Use')
        ws.sheet_view.showGridLines = False
        ws.column_dimensions['A'].width = 3
        ws.column_dimensions['B'].width = 22
        ws.column_dimensions['C'].width = 70

        self._title_cell(ws, 1, 1, 'wfmplan  |  Methodology & User Guide', span=3, bg=DARK_BLUE, size=14)
        ws.row_dimensions[1].height = 35

        content = [
            ('', '', ''),
            ('METHODOLOGY', '', ''),
            ('', 'Erlang-C Model', 'Models incoming contacts as a Poisson process (M/M/c queue). '
                                   'Assumes random, memoryless arrivals and exponential service times.'),
            ('', 'Traffic Intensity (ρ)', 'ρ = (Volume / Interval) × AHT  '
                                          'Measured in Erlangs. Represents average concurrent agents busy.'),
            ('', 'P(Wait) / Erlang-C', 'C(c,ρ) – probability that an arriving contact must queue. '
                                        'Decreases as more agents are added.'),
            ('', 'ASA', 'Average Speed of Answer = C(c,ρ) × AHT / (c – ρ)  [seconds]'),
            ('', 'SLA', 'SLA = 1 – C(c,ρ) × e^(–(c–ρ)×ST/AHT)  '
                         'Fraction of contacts answered within service target ST seconds.'),
            ('', 'Shrinkage', 'agent_req_shrink = ceil(agent_req / (1 – shrink)). '
                               'Accounts for breaks, training, absence, etc.'),
            ('', 'Max Occupancy', 'If ρ/c exceeds max_occupancy, c is increased until the cap is met. '
                                   'Prevents agent burn-out (recommended: 0.85–0.90 for voice).'),
            ('', '', ''),
            ('INPUT COLUMNS', '', ''),
            ('', 'exp_vol', 'Expected contact volume for the interval  (required)'),
            ('', 'exp_aht', 'Average Handling Time in seconds  (required)'),
            ('', 'interval_start', 'ISO datetime – start of the planning interval  (required)'),
            ('', 'interval_end', 'ISO datetime – end of the planning interval  (required)'),
            ('', 'max_occupancy', 'Per-row occupancy cap override  (optional, e.g. 0.85)'),
            ('', 'shrink', 'Per-row shrinkage override  (optional, e.g. 0.15)'),
            ('', 'asa / sla / st', 'Per-row target overrides  (optional)'),
            ('', '', ''),
            ('OUTPUT COLUMNS', '', ''),
            ('', 'agent_req', 'Minimum agents to hit SLA/ASA target'),
            ('', 'agent_req_shrink', 'Scheduled headcount after shrinkage'),
            ('', 'traffic_intensity', 'Erlangs (ρ)'),
            ('', 'pred_occupancy', 'Fraction of time agents are busy  (0–1)'),
            ('', 'prob_waiting', 'Erlang-C probability a contact must wait'),
            ('', 'pred_asa', 'Predicted Average Speed of Answer (seconds)'),
            ('', 'pred_sla_pct', 'Predicted SLA percentage  (if st provided)'),
            ('', '', ''),
            ('SHIFTS', '', ''),
            ('', 'Default shifts', 'Early (06:00), Morning (08:00), Mid (10:00), '
                                    'Afternoon (12:00), Late (14:00), Evening (16:00), Night (22:00)'),
            ('', 'Custom shifts', 'Pass a list of ShiftTemplate objects to WeeklyScheduler(shift_templates=…)'),
            ('', 'Scheduling logic', 'Greedy covering: repeatedly assigns one agent to the shift that '
                                      'reduces the most unmet demand across its covered intervals.'),
        ]

        for r, (section, label, text) in enumerate(content, 3):
            ws.row_dimensions[r].height = 18
            if section:
                c = ws.cell(row=r, column=2, value=section)
                c.font = _font(bold=True, size=12, color=WHITE)
                c.fill = _fill(MID_BLUE)
                c.alignment = _left()
                ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
            else:
                lc = ws.cell(row=r, column=2, value=label)
                lc.font = _font(bold=bool(label), size=10)
                lc.alignment = _left()
                tc = ws.cell(row=r, column=3, value=text)
                tc.font = _font(size=10)
                tc.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                ws.row_dimensions[r].height = 28 if len(text) > 80 else 18
